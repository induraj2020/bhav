import streamlit as st
import pandas as pd
import io
import ast
import warnings
from datetime import datetime
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.filterwarnings("ignore")
pd.set_option('display.max_rows', 5000)
pd.set_option('display.max_columns', 1000)
import sys
import re
from datetime import datetime, timedelta
from pytz import timezone, all_timezones
import time
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.styles import Font
pd.set_option('display.max_rows', 5000)
pd.set_option('display.max_columns', 1000)
from openpyxl import load_workbook
# from utils import *
from io import BytesIO

def find_eod_oichange(l_df):
    # total value of all the open contracts at the end of the day.
    # total value of the change in open contracts
    l_df['EOD'] = l_df['OpnIntrst']*l_df['ClsPric']
    l_df['OICHANGE_Everyday'] = l_df['ChngInOpnIntrst']*l_df['ClsPric']
    return l_df

def filter_by_ce_pe(l_df, date):
    l_df['XpryDt'] = pd.to_datetime(l_df['XpryDt'])
    expirty_df = l_df[l_df['XpryDt']== pd.to_datetime(date)]
    expirty_df = expirty_df.reset_index(drop=True)
    # display(HTML(expirty_df.to_html()))
    expirty_df['TradDt'] =  pd.to_datetime(expirty_df.loc[0]['TradDt']).date()
    expirty_df['XpryDt'] = pd.to_datetime(expirty_df.loc[0]['XpryDt']).date()
    expirty_CE = expirty_df[expirty_df['OptnTp']== 'CE'].reset_index(drop=True)
    expirty_CE = find_eod_oichange(expirty_CE)
    expirty_PE = expirty_df[expirty_df['OptnTp']== 'PE'].reset_index(drop=True)
    expirty_PE = find_eod_oichange(expirty_PE)
    expirty_CE = expirty_CE.sort_values(['OICHANGE_Everyday'], ascending=False).reset_index(drop=True)
    expirty_PE = expirty_PE.sort_values(['OICHANGE_Everyday'], ascending=False).reset_index(drop=True)
    expirty_CE = expirty_CE[['TradDt', 'XpryDt', 'OpnIntrst', 'ChngInOpnIntrst', 'OptnTp', 'ClsPric', 'EOD', 'OICHANGE_Everyday','StrkPric']]
    expirty_PE = expirty_PE[['TradDt', 'XpryDt', 'OpnIntrst', 'ChngInOpnIntrst', 'OptnTp', 'ClsPric', 'EOD', 'OICHANGE_Everyday','StrkPric']]
    return (expirty_CE, expirty_PE)
    
def format_number(num):
    neg = False
    if num<0:
        neg = True
        num = abs(num)
    if num >= 1_00_00_00_000: 
        final_num =  f"{num/1_00_00_000:.2f} cr"
    elif num >= 1_00_00_000:  
        final_num = f"{num/1_00_00_000:.2f} cr"
    elif num >= 1_00_000:
        final_num = f"{num/1_00_000:.2f} lk"
    elif num >= 1_000:  
        final_num = f"{num/1_000:.2f} th"
    elif num <= 1_000:  
        final_num= f"{num/1_000:.2f} th"
    if neg==True:
        return '-'+final_num
    else:
        return final_num

def format_number_in_crores(num):
    neg = False
    if num<0:
        neg = True
        num = abs(num)
    final_num =  round(num/10000000,3)
    if neg==True:
        return -abs(final_num)
    else:
        return final_num
    
def find_changes(l_df_ce, l_df_pe, spot):
    dic = {}
    dic['Date'] = pd.to_datetime(l_df_ce.loc[0]['TradDt']).date()
    dic['Expiry'] = pd.to_datetime(l_df_ce.loc[0]['XpryDt']).date()
    dic['Spot'] = spot
    ce_eod = round(sum(l_df_ce['EOD']),2)
    ce_change = round(sum(l_df_ce['OICHANGE_Everyday']),2)  
    dic['CE_EOD'] = format_number_in_crores(ce_eod)
    dic['CE_CHANGE'] = format_number_in_crores(ce_change)
    dic['CE_%CHANGE'] = round(ce_change/ce_eod,4)*100
    
    
    itm_df = l_df_ce[l_df_ce['StrkPric']<spot]
    itm_eod = sum(itm_df['EOD'])
    
    dic['ITM_CE_EOD'] = format_number_in_crores(itm_eod)
    dic['ITM_CE_%EOD'] = round(itm_eod/ce_eod,4)*100
    
    itm_oi_change = round(sum(itm_df['OICHANGE_Everyday']),4)
    dic['ITM_CE_CHANGE'] = format_number_in_crores(itm_oi_change)
    dic['ITM_CE_%CHANGE'] = round(itm_oi_change/itm_eod,4)*100

    pe_eod = round(sum(l_df_pe['EOD']),2)
    pe_change = round(sum(l_df_pe['OICHANGE_Everyday']),2)  

    dic['PE_EOD'] = format_number_in_crores(pe_eod)
    dic['PE_CHANGE'] = format_number_in_crores(pe_change)
    dic['PE_%CHANGE'] = round(pe_change/pe_eod,4)*100
    
    itm_df = l_df_pe[l_df_pe['StrkPric']>spot]
    itm_eod = sum(itm_df['EOD'])
    
    dic['ITM_PE_EOD'] =  format_number_in_crores(itm_eod)
    dic['ITM_PE_%EOD'] = round(itm_eod/pe_eod,4)*100
    itm_oi_change = round(sum(itm_df['OICHANGE_Everyday']),2)
    dic['ITM_PE_CHANGE'] = format_number_in_crores(itm_oi_change)
    dic['ITM_PE_%CHANGE'] = round(itm_oi_change/itm_eod,4)*100
    for key in dic:
        dic[key] = [dic[key]]
    l_df = pd.DataFrame(dic)
    l_df['Date'] = pd.to_datetime(l_df['Date']).dt.date
    l_df['Expiry'] = pd.to_datetime(l_df['Expiry']).dt.date
    return l_df

def save_df_as_excel(file_loc, final_df, sheet_name):
    book = load_workbook(file_loc)
    if sheet_name in book.sheetnames:
        existing_df = pd.read_excel(file_loc, sheet_name=sheet_name)
        if "Date" in existing_df.columns:
            existing_df['Date'] = pd.to_datetime(existing_df['Date']).dt.date
        if "Expiry" in existing_df.columns:
            existing_df['Expiry'] = pd.to_datetime(existing_df['Date']).dt.date    
        final_df = pd.concat([existing_df, final_df], ignore_index=True)
    return final_df